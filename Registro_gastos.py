from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QFormLayout, \
    QLineEdit, QPushButton, QMessageBox, QComboBox, QStackedWidget, QTableWidget, \
    QTableWidgetItem, QHeaderView, QDateEdit, QMenuBar, QHBoxLayout
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QGuiApplication
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
import sqlite3
import sys
import pandas as pd
from openpyxl import Workbook
from fpdf import FPDF

class RegistroGastos(QWidget):
    def __init__(self, grafico_widget):
        super().__init__()
        self.grafico_widget = grafico_widget  # Pasar el widget del gráfico
        self.initUI()
        self.create_database()
        self.load_data()

    def initUI(self):
        self.layout = QVBoxLayout()

        # Formulario para introducir datos
        self.form_layout = QFormLayout()

        self.descripcion_input = QLineEdit()
        self.montopresupuestado_input = QLineEdit()
        self.montoreal_input = QLineEdit()
        self.categoria_input = QComboBox()

        # QDateEdit para seleccionar la fecha
        self.fecha_input = QDateEdit()
        self.fecha_input.setCalendarPopup(True)  # Habilitar el selector de calendario
        self.fecha_input.setDate(QDate.currentDate())  # Fecha predeterminada: hoy
        # Campo adicional para "Otra Categoría"
        self.categoria_extra_input = QLineEdit()
        self.categoria_extra_input.setPlaceholderText("Escribe la categoría aquí")
        self.categoria_extra_input.setVisible(False)  # Oculto por defecto
        # Conectar el cambio de selección en el QComboBox
        self.categoria_input.currentIndexChanged.connect(self.toggle_categoria_extra)

        # Agregar opciones al QComboBox
        self.categoria_input.addItems(['Gasto fijo', 'Personal/ocio', 'Préstamos/tarjetas', 'Ahorro', 'Imprevisto','Otra'])

        self.form_layout.addRow('Descripción:', self.descripcion_input)
        self.form_layout.addRow('Gasto presupuestado:', self.montopresupuestado_input)
        self.form_layout.addRow('Monto Real:', self.montoreal_input)
        self.form_layout.addRow('Categoría:', self.categoria_input)
        self.form_layout.addRow('Otra Categoría:', self.categoria_extra_input)
        self.form_layout.addRow('Fecha:', self.fecha_input)

        self.layout.addLayout(self.form_layout)

        self.submit_button = QPushButton('Registrar Gasto')
        self.submit_button.setStyleSheet("background-color: #4CAF50; color: white; border-radius: 5px; padding: 10px;")
        self.submit_button.clicked.connect(self.submit_data)
        self.layout.addWidget(self.submit_button)

        self.modify_button = QPushButton('Modificar Gasto')
        self.modify_button.setStyleSheet("background-color: #FF9800; color: white; border-radius: 5px; padding: 10px;")
        self.modify_button.clicked.connect(self.modify_data)
        self.layout.addWidget(self.modify_button)

        self.delete_button = QPushButton('Eliminar Gasto')
        self.delete_button.setStyleSheet("background-color: #F44336; color: white; border-radius: 5px; padding: 10px;")
        self.delete_button.clicked.connect(self.delete_data)
        self.layout.addWidget(self.delete_button)

        # Tabla para mostrar los datos
        self.gastos_table = QTableWidget()
        self.gastos_table.setColumnCount(5)
        self.gastos_table.setHorizontalHeaderLabels(['Descripción', 'Presupuestado', 'Real', 'Categoría', 'Fecha'])
        self.gastos_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.gastos_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.gastos_table.setSelectionMode(QTableWidget.SingleSelection)
        self.gastos_table.setStyleSheet("background-color: #FFFFFF; border: 1px solid #DDDDDD;")
        self.layout.addWidget(self.gastos_table)

        self.setLayout(self.layout)

    def toggle_categoria_extra(self):
        """Muestra u oculta el campo para categoría extra según la selección."""
        if self.categoria_input.currentText() == "Otra":
            self.categoria_extra_input.setVisible(True)
        else:
            self.categoria_extra_input.setVisible(False)
            self.categoria_extra_input.clear()

    def create_database(self):
        conn = sqlite3.connect('gastos.db')
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS gastos (
            gastosid INTEGER PRIMARY KEY AUTOINCREMENT,
            montopresupuestado REAL NOT NULL,
            descripcion TEXT NOT NULL,
            montoreal REAL NOT NULL,
            categoria TEXT NOT NULL,
            fecha DATE NOT NULL
        )''')
        conn.commit()
        conn.close()

    def submit_data(self):
        descripcion = self.descripcion_input.text()
        montopresupuestado = self.montopresupuestado_input.text()
        montoreal = self.montoreal_input.text()
        categoria = self.categoria_input.currentText()
        fecha = self.fecha_input.date().toString("yyyy-MM-dd")  # Convertir la fecha seleccionada a formato string

        # Usar el valor de categoría extra si está seleccionada
        if categoria == "Otra":
            categoria = self.categoria_extra_input.text().strip()

        # Validación de campos vacíos
        if not descripcion or not montopresupuestado or not montoreal or not fecha:
            QMessageBox.warning(self, 'Error', 'Todos los campos son obligatorios.')
            return

        # Validación de los montos
        try:
            montopresupuestado = float(montopresupuestado)
            montoreal = float(montoreal)
        except ValueError:
            QMessageBox.warning(self, 'Error', 'El monto presupuestado y el monto real deben ser números válidos.')
            return

        conn = sqlite3.connect('gastos.db')
        cursor = conn.cursor()
        cursor.execute('''INSERT INTO gastos (montopresupuestado, descripcion, montoreal, categoria, fecha)
                          VALUES (?, ?, ?, ?, ?)''', (montopresupuestado, descripcion, montoreal, categoria, fecha))
        conn.commit()
        conn.close()

        self.montopresupuestado_input.clear()
        self.descripcion_input.clear()
        self.montoreal_input.clear()
        self.fecha_input.setDate(QDate.currentDate())  # Restablecer la fecha al valor actual

        QMessageBox.information(self, 'Éxito', 'Gasto registrado exitosamente.')
        self.load_data()

        # Actualiza el gráfico inmediatamente
        self.grafico_widget.update_chart()

    def modify_data(self):
        selected_row = self.gastos_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, 'Error', 'Por favor, selecciona un gasto para modificar.')
            return

        gasto_id = self.gastos_table.item(selected_row, 0).data(Qt.UserRole)

        conn = sqlite3.connect('gastos.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM gastos WHERE gastosid = ?", (gasto_id,))
        row = cursor.fetchone()
        conn.close()

        # Asegurarse de convertir los valores float a cadena
        self.descripcion_input.setText(row[2])
        self.montopresupuestado_input.setText(str(row[1]))  # Convertir a string
        self.montoreal_input.setText(str(row[3]))  # Convertir a string
        self.categoria_input.setCurrentText(row[4])
        self.fecha_input.setDate(QDate.fromString(row[5], "yyyy-MM-dd"))  # Convertir la fecha de la base de datos a formato QDate

        # Cambiar texto del botón y asociar una función de guardar cambios
        self.submit_button.setText("Guardar Cambios")
        self.submit_button.clicked.disconnect()  # Desconectar la función anterior
        self.submit_button.clicked.connect(lambda: self.save_changes(gasto_id))  # Conectar a la nueva función

    def save_changes(self, gasto_id):
        descripcion = self.descripcion_input.text()
        montopresupuestado = self.montopresupuestado_input.text()
        montoreal = self.montoreal_input.text()
        categoria = self.categoria_input.currentText()
        fecha = self.fecha_input.date().toString("yyyy-MM-dd")  # Convertir la fecha seleccionada a formato string
        # Usar el valor de categoría extra si está seleccionada
        if categoria == "Otra":
            categoria = self.categoria_extra_input.text().strip()

        # Validación de campos vacíos
        if not descripcion or not montopresupuestado or not montoreal or not fecha:
            QMessageBox.warning(self, 'Error', 'Todos los campos son obligatorios.')
            return

        # Validación de los montos
        try:
            montopresupuestado = float(montopresupuestado)
            montoreal = float(montoreal)
        except ValueError:
            QMessageBox.warning(self, 'Error', 'El monto presupuestado y el monto real deben ser números válidos.')
            return

        conn = sqlite3.connect('gastos.db')
        cursor = conn.cursor()
        cursor.execute('''UPDATE gastos SET montopresupuestado = ?, descripcion = ?, montoreal = ?, categoria = ?, fecha = ? 
                          WHERE gastosid = ?''', (montopresupuestado, descripcion, montoreal, categoria, fecha, gasto_id))
        conn.commit()
        conn.close()

        self.montopresupuestado_input.clear()
        self.descripcion_input.clear()
        self.montoreal_input.clear()
        self.fecha_input.setDate(QDate.currentDate())  # Restablecer la fecha al valor actual

        QMessageBox.information(self, 'Éxito', 'Gasto modificado exitosamente.')
        self.load_data()

        # Actualiza el gráfico inmediatamente
        self.grafico_widget.update_chart()

        self.submit_button.setText("Registrar Gasto")
        self.submit_button.clicked.disconnect()
        self.submit_button.clicked.connect(self.submit_data)  # Volver a la función original

    def delete_data(self):
        selected_row = self.gastos_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, 'Error', 'Por favor, selecciona un gasto para eliminar.')
            return

        gasto_id = self.gastos_table.item(selected_row, 0).data(Qt.UserRole)

        confirm = QMessageBox.question(self, 'Confirmar eliminación', '¿Estás seguro de que quieres eliminar este gasto?',
                                       QMessageBox.Yes | QMessageBox.No)
        if confirm == QMessageBox.Yes:
            conn = sqlite3.connect('gastos.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM gastos WHERE gastosid = ?", (gasto_id,))
            conn.commit()
            conn.close()

            self.load_data()

            # Actualiza el gráfico inmediatamente
            self.grafico_widget.update_chart()

            QMessageBox.information(self, 'Éxito', 'Gasto eliminado exitosamente.')

    def load_data(self):
        conn = sqlite3.connect('gastos.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM gastos")
        rows = cursor.fetchall()
        conn.close()

        self.gastos_table.setRowCount(0)
        for row in rows:
            row_position = self.gastos_table.rowCount()
            self.gastos_table.insertRow(row_position)
            self.gastos_table.setItem(row_position, 0, QTableWidgetItem(str(row[2])))  # gastosid
            self.gastos_table.setItem(row_position, 1, QTableWidgetItem(str(row[1])))  # montopresupuestado
            self.gastos_table.setItem(row_position, 2, QTableWidgetItem(str(row[3])))  # montoreal
            self.gastos_table.setItem(row_position, 3, QTableWidgetItem(row[4]))  # categoria
            self.gastos_table.setItem(row_position, 4, QTableWidgetItem(row[5]))  # fecha
            self.gastos_table.item(row_position, 0).setData(Qt.UserRole, row[0])  # guardar el id del gasto

    def compare_data(self):
        # Función para comparar los montos presupuestados y reales usando pandas
        conn = sqlite3.connect('gastos.db')
        query = "SELECT descripcion, montopresupuestado, montoreal, categoria, fecha FROM gastos"
        df = pd.read_sql(query, conn)
        conn.close()

        # Comparar los montos presupuestados vs. reales
        df['diferencia'] = df['montopresupuestado'] - df['montoreal']
        return df

    def export_to_excel(self, df):
        # Función para exportar los datos a un archivo Excel usando OpenPyXL
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Gastos"
        
        # Agregar encabezados
        ws.append(['Descripción', 'Monto Presupuestado', 'Monto Real', 'Categoría', 'Fecha', 'Diferencia'])

        for row in df.itertuples():
            ws.append([row.descripcion, row.montopresupuestado, row.montoreal, row.categoria, row.fecha, row.diferencia])

        # Guardar el archivo Excel
        wb.save("historial_gastos.xlsx")
        QMessageBox.information(self, 'Éxito', 'Datos exportados a Excel exitosamente.')

    def export_to_pdf(self, df):
        # Función para exportar los datos a un archivo PDF usando FPDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)  # Permitir saltos de página automáticos
        pdf.add_page()
        pdf.set_font("Arial", size=10)

        pdf.cell(200, 10, txt="Historial de Gastos", ln=True, align="C")
        pdf.ln(10)

        # Ajuste dinámico de anchos de columna
        column_widths = {
            'Descripción': 40,
            'Monto Presupuestado': 35,
            'Monto Real': 30,
            'Categoría': 30,
            'Fecha': 30,
            'Diferencia': 30
        }
    
         # Crear encabezados en PDF
        for col_name in column_widths:
           pdf.cell(column_widths[col_name], 10, col_name, border=5, align='C')
        pdf.ln()

        # Agregar filas
        for row in df.itertuples():
            pdf.cell(column_widths['Descripción'], 10, str(row.descripcion)[:40], border=1)  # Limitar texto a 40 caracteres
            pdf.cell(column_widths['Monto Presupuestado'], 10, str(row.montopresupuestado), border=1, align='C')
            pdf.cell(column_widths['Monto Real'], 10, str(row.montoreal), border=1, align='C')
            pdf.cell(column_widths['Categoría'], 10, str(row.categoria)[:30], border=1)  # Limitar texto a 30 caracteres
            pdf.cell(column_widths['Fecha'], 10, str(row.fecha), border=1, align='C')
            pdf.cell(column_widths['Diferencia'], 10, str(row.diferencia), border=1, align='C')
            pdf.ln()

        # Guardar el archivo PDF
        pdf.output("historial_gastos.pdf")
        QMessageBox.information(self, 'Éxito', 'Datos exportados a PDF exitosamente.')


    def show_comparison(self):
        df = self.compare_data()

        # Mostrar la comparación en un cuadro de mensaje
        comparison_text = df[['descripcion', 'montopresupuestado', 'montoreal', 'diferencia']].to_string(index=False)
        QMessageBox.information(self, 'Comparación de Gastos', comparison_text)

        return df

class GraficoPorCategoriaWidget(QWidget):
    def __init__(self):
        super().__init__()

        # Layout del widget
        self.layout = QVBoxLayout()

        # Selector de fechas
        self.date_layout = QHBoxLayout()
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate().addMonths(-1))  # Fecha de inicio predeterminada
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate())  # Fecha de fin predeterminada

        self.generate_button = QPushButton('Generar Gráfico')
        self.generate_button.clicked.connect(self.update_chart)

        self.date_layout.addWidget(self.start_date)
        self.date_layout.addWidget(self.end_date)
        self.date_layout.addWidget(self.generate_button)

        self.layout.addLayout(self.date_layout)

        # Configuración del gráfico
        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)
        self.layout.addWidget(self.canvas)

        self.setLayout(self.layout)

    def update_chart(self):
        # Obtener las fechas seleccionadas
        start_date = self.start_date.date().toString("yyyy-MM-dd")
        end_date = self.end_date.date().toString("yyyy-MM-dd")

        # Conectar a la base de datos
        conn = sqlite3.connect('gastos.db')
        query = '''
        SELECT categoria, SUM(montoreal) as total 
        FROM gastos 
        WHERE fecha BETWEEN ? AND ? 
        GROUP BY categoria
        '''
        df = pd.read_sql_query(query, conn, params=(start_date, end_date))
        conn.close()

        # Limpiar el gráfico actual
        self.ax.clear()

        # Si hay datos, genera el gráfico
        if not df.empty:
            self.ax.pie(
                df['total'], 
                labels=df['categoria'], 
                autopct='%1.1f%%', 
                startangle=140, 
                textprops={'fontsize': 10}
            )
            self.ax.set_title('Gastos por Categoría')
        else:
            self.ax.text(0.5, 0.5, 'No hay datos para este rango de fechas', 
                         horizontalalignment='center', verticalalignment='center', fontsize=12)

        # Actualizar el canvas
        self.canvas.draw()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestión de Gastos")
        self.setGeometry(100, 100, 800, 600)

        self.central_widget = QStackedWidget()
        self.setCentralWidget(self.central_widget)

        self.grafico_widget = GraficoPorCategoriaWidget()  
        self.registro_gastos = RegistroGastos(self.grafico_widget)
        self.central_widget.addWidget(self.registro_gastos)
        self.central_widget.addWidget(self.grafico_widget)

        self.init_menu()

    def init_menu(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu('Menú')

        registro_action = file_menu.addAction('Registro de Gastos')
        registro_action.triggered.connect(self.show_registro)

        grafico_action = file_menu.addAction('Ver Gráficos')
        grafico_action.triggered.connect(self.show_grafico)

        # Opción para ver la comparación de gastos
        compare_action = file_menu.addAction('Comparar Gastos')
        compare_action.triggered.connect(self.show_comparison)

        # Opción para exportar a Excel
        export_excel_action = file_menu.addAction('Exportar a Excel')
        export_excel_action.triggered.connect(self.export_to_excel)

        # Opción para exportar a PDF
        export_pdf_action = file_menu.addAction('Exportar a PDF')
        export_pdf_action.triggered.connect(self.export_to_pdf)

    def show_registro(self):
        self.central_widget.setCurrentWidget(self.registro_gastos)

    def show_grafico(self):
        self.central_widget.setCurrentWidget(self.grafico_widget)

    def show_comparison(self):
        df = self.registro_gastos.show_comparison()

    def export_to_excel(self):
        df = self.registro_gastos.compare_data()
        self.registro_gastos.export_to_excel(df)

    def export_to_pdf(self):
        df = self.registro_gastos.compare_data()
        self.registro_gastos.export_to_pdf(df)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
